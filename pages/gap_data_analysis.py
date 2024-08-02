import streamlit as st 
import pandas as pd
import numpy as np
from io import BytesIO
import os
import datetime
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import numbers
import plotly.graph_objects as go
import snowflake.connector

# Import custom modules
from login import  login  #login_form
from menu import menu_with_redirect
from util import apply_custom_style, add_logo, get_logo_url, get_logo_path, render_home_sidebar, style_metric_cards, fetch_supplier_names, render_gap_analysis_sidebar
from snowflake_connection import create_gap_report, get_snowflake_connection, execute_query_and_close_connection, get_snowflake_toml, validate_toml_info, fetch_and_store_toml_info, fetch_chain_schematic_data

st.set_page_config(layout="wide")

# Redirect to Chainlink_Main.py if not logged in, otherwise show the navigation menu
render_gap_analysis_sidebar()
menu_with_redirect()

def format_sales_report(workbook):
    # Delete all sheets except SALES REPORT
    for sheet_name in workbook.sheetnames:
        if sheet_name != 'SALES REPORT':
            workbook.remove(workbook[sheet_name])

    # Select the SALES REPORT sheet
    ws = workbook['SALES REPORT']

    # Delete row 2
    ws.delete_rows(2)

    # Delete column H
    ws.delete_cols(8)

    # Remove all hyphens from column F
    for cell in ws['F']:
        if cell.value is not None:
            cell.value = str(cell.value).replace('-', '')

    # Create a new column for store name
    ws.insert_cols(2)
    ws.cell(row=1, column=2, value='STORE NAME')

    # Copy values before the # to store name column
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell_offset = ws.cell(row=cell.row, column=2)
            cell_value = str(cell.value)

            # Check if '#' is in the cell value
            if '#' in cell_value:
                store_name = cell_value.split('#')[0].replace("'", "")
            else:
                # If '#' is not present, keep the entire cell value
                store_name = cell_value.replace("'", "")

            cell_offset.value = store_name

    # Remove column C
    ws.delete_cols(3)

    # Replace all commas with spaces in column B
    for cell in ws['B']:
        if cell.value is not None and isinstance(cell.value, str):
            cell.value = cell.value.replace(',', ' ')

    # Remove all 's in column B
    for cell in ws['B']:
        if cell.value is not None and isinstance(cell.value, str):
            cell.value = cell.value.replace(" 's", "")

    # Replace all commas with spaces in column E
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str):
                cell.value = cell.value.replace(',', ' ')

    # Replace all single quote with spaces in column E
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str):
                cell.value = cell.value.replace("'", ' ')

    # Remove all commas from column C
    for cell in ws['C']:
        if cell.value is not None:
            cell.value = str(cell.value).replace(',', ' ')

    # Remove all Is Null from column F
    for cell in ws['F']:
        if cell.value is not None:
            cell.value = str(cell.value).replace('Is Null', '0')

    # Format column G as number with no decimals
    for cell in ws['G'][1:]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = numbers.FORMAT_NUMBER
        elif isinstance(cell.value, str):
            cell.number_format = numbers.FORMAT_NUMBER
            try:
                cell.value = float(cell.value.replace(",", ""))
            except ValueError:
                pass

    return workbook

# Upload the workbook
uploaded_file = st.file_uploader(":red[Upload freshly ran sales table from your application]", type=["xlsx", "xls"])

if uploaded_file is not None:
    # Load the workbook
    workbook = load_workbook(uploaded_file)

    # Show the Reformat button
    if st.button("Reformat"):
        # Format the sales report
        new_workbook = format_sales_report(workbook)

        # Download the formatted file
        new_filename = 'formatted_' + uploaded_file.name
        stream = BytesIO()
        new_workbook.save(stream)
        stream.seek(0)
        st.download_button(label="Download formatted file", data=stream.read(), file_name=new_filename, mime='application/vnd.ms-excel')
        
#========================================================================================================================================
# Function to write sales report data to snowflake
# #======================================================================================================================================        

def write_salesreport_to_snowflake(df):
    try:
        # Read Excel file into pandas DataFrame
       # df = pd.read_excel(uploaded_file)

        # Replace NaN values with "NULL"
        df.fillna(value="NULL", inplace=True)

        toml_info = st.session_state.get('toml_info')
        if not toml_info:
            st.error("TOML information is not available. Please check the tenant ID and try again.")
            return

        # Create a connection to Snowflake
        conn_toml = get_snowflake_toml(toml_info)

        # Create a cursor object
        cursor = conn_toml.cursor()

        # Get the current timestamp
        current_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Write DataFrame to Snowflake
        sql_query = f"""
        CREATE OR REPLACE TABLE SALES_REPORT AS
        SELECT 
            CAST(STORE_NUMBER AS NUMBER) AS STORE_NUMBER, 
            CAST(TRIM(STORE_NAME) AS VARCHAR) AS STORE_NAME, 
            CAST(ADDRESS AS VARCHAR) AS ADDRESS, 
            CAST(SALESPERSON AS VARCHAR) AS SALESPERSON, 
            CAST(PRODUCT_NAME AS VARCHAR) AS PRODUCT_NAME, 
            CAST(UPC AS NUMERIC) AS UPC, 
            CAST(PURCHASED_YES_NO AS NUMERIC) AS PURCHASED_YES_NO,
            CAST('{current_timestamp}' AS TIMESTAMP) AS LAST_UPLOAD_DATE
        FROM (VALUES {', '.join([str(tuple(row)) for row in df.values])}) 
        AS tmp(STORE_NUMBER, STORE_NAME, ADDRESS, SALESPERSON, PRODUCT_NAME, UPC, PURCHASED_YES_NO);
        """

        # Execute the SQL query
        cursor.execute(sql_query)
        conn_toml.commit()

        st.success("Data has been imported into Snowflake table SALES_REPORT!")

    except snowflake.connector.errors.Error as e:
        st.error(f"Error writing to Snowflake: {str(e)}")
    except Exception as e:
        st.error(f"An unexpected error occurred: {str(e)}")
    finally:
        # Ensure that the cursor and connection are closed properly
        if 'cursor' in locals():
            cursor.close()
        if 'conn_toml' in locals():
            conn_toml.close()


        
#========================================================================================================================================
# END Function to write sales report data to snowflake
# #======================================================================================================================================        



# Create uploader for formatted sales report create dataframe and call write to snowflake function
uploaded_file = st.file_uploader(":red[UPLOAD CURRENT SALES REPORT AFTER IT HAS BEEN FORMATED]", type=["xlsx"])

# Check if file was uploaded
if uploaded_file:
    # Read Excel file into pandas DataFrame
    df = pd.read_excel(uploaded_file)
    #print(df.columns)
    # Display DataFrame in Streamlit
    #st.dataframe(df)

    # Write DataFrame to Snowflake on button click
    if st.button("Import into Snowflake"):
        with st.spinner('Uploading Sales Report Data to Snowflake ...'):
            write_salesreport_to_snowflake(df)

# Function to create and display the gap analysis bar chart
def gap_analysis_bar_chart():
    toml_info = st.session_state.get('toml_info')
    if not toml_info:
        st.error("TOML information is not available. Please check the tenant ID and try again.")
        return

    # Create a connection to Snowflake
    conn_toml = get_snowflake_toml(toml_info)

    # Create a cursor object
    cursor = conn_toml.cursor()

    # Retrieve data from your view
    query = """
        SELECT 
            SUM("In_Schematic") AS total_in_schematic, 
            SUM("PURCHASED_YES_NO") AS purchased, 
            SUM("PURCHASED_YES_NO") / COUNT(*) AS purchased_percentage 
        FROM GAP_REPORT;
    """
    df = pd.read_sql(query, conn_toml)

    # Check if the DataFrame is not empty and has the expected columns
    expected_columns = ['TOTAL_IN_SCHEMATIC', 'PURCHASED', 'PURCHASED_PERCENTAGE']
    if df.empty:
        st.error("No data retrieved from GAP_REPORT.")
        return

    if not all(col in df.columns for col in expected_columns):
        st.error(f"Unexpected columns in the retrieved data. Expected columns: {expected_columns}, but got: {df.columns.tolist()}")
        return

    # Format the 'PURCHASED_PERCENTAGE' column as a percentage with 2 decimal places
    df['PURCHASED_PERCENTAGE'] = (df['PURCHASED'] / df['TOTAL_IN_SCHEMATIC'] * 100).map('{:.2f}%'.format)

    # Create the bar chart
    fig = go.Figure(data=[
        go.Bar(
            x=['Total in Schematic', 'Purchased', 'Purchased Percentage'], 
            y=[df['TOTAL_IN_SCHEMATIC'].iloc[0], df['PURCHASED'].iloc[0], df['PURCHASED_PERCENTAGE'].iloc[0]], 
            text=[df['TOTAL_IN_SCHEMATIC'].iloc[0], df['PURCHASED'].iloc[0], df['PURCHASED_PERCENTAGE'].iloc[0]], 
            textposition='auto', 
            marker=dict(color=['#1f77b4', '#ff7f0e', '#2ca02c'])
        )
    ])

    # Set the axis labels and plot title
    fig.update_layout(
        xaxis_title='',
        yaxis_title='Number of Items',
        title='Total Items in Schematic vs. Purchased Items',
        plot_bgcolor='#B3D7ED',  # Set the background color
        paper_bgcolor='#F8F2EB' # Set the paper (border) color
    )
  
    # Add a border to the chart
    fig.update_traces(
        marker_line_width=1.5,
        marker_line_color='black'
    )

    # Customize the bar chart colors
    colors = ['#1f77b4', '#ff7f0e', '#2ca02c']
    for i in range(len(fig.data)):
        fig.data[i].marker.color = colors[i]
        fig.data[i].marker.line.width = 1.5
        fig.data[i].marker.line.color = 'black'

    # Row A
    col1, col2 = st.columns(2)

    with col1:
        container = st.container()
        with container:
            st.plotly_chart(fig, use_container_width=True)

    with col2:
        container = st.container()
        with container:
            st.plotly_chart(fig, use_container_width=True)



def main():
    
    if 'authenticated' not in st.session_state or not st.session_state['authenticated']:
        login()
    else:
        tenant_id = st.session_state.get('tenant_id')
        if tenant_id and not st.session_state.get('toml_info'):
            if not fetch_and_store_toml_info(tenant_id):
                st.error("Failed to retrieve or validate configuration.")
                return

        if 'toml_info' in st.session_state:
            # Call the function to display the gap analysis bar chart
            gap_analysis_bar_chart()
            #st.write(st.session_state.toml_info)
        
            

if __name__ == "__main__":
    main()
